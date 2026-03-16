# app/carga_masiva.py
"""
Módulo de carga masiva.
Acepta un archivo Excel o TXT (pipes |) con la malla CAC y valida cada fila.
Devuelve un Excel con los errores encontrados y un resumen.

Formato Excel esperado:
  Columna A = V1, B = V2, ..., hasta la columna correspondiente a V134
  La primera fila puede ser cabecera (se detecta automáticamente).

Formato TXT esperado:
  Cada línea es un registro, campos separados por | (pipe).
  168 campos en el orden del anexo técnico de la Resolución 0247/2014.
"""

import io
import re
from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd

# Mapa posición (1-based) → clave del schema
# 134 variables principales + subcampos especiales de V46 y V114
POSICION_A_VAR: dict[int, str] = {
    1: "V1", 2: "V2", 3: "V3", 4: "V4", 5: "V5", 6: "V6", 7: "V7",
    8: "V8", 9: "V9", 10: "V10", 11: "V11", 12: "V12", 13: "V13",
    14: "V14", 15: "V15", 16: "V16", 17: "V17", 18: "V18", 19: "V19",
    20: "V20", 21: "V21", 22: "V22", 23: "V23", 24: "V24", 25: "V25",
    26: "V26", 27: "V27", 28: "V28", 29: "V29", 30: "V30", 31: "V31",
    32: "V32", 33: "V33", 34: "V34", 35: "V35", 36: "V36", 37: "V37",
    38: "V38", 39: "V39", 40: "V40", 41: "V41", 42: "V42", 43: "V43",
    44: "V44", 45: "V45", 46: "V46", 47: "V46_1", 48: "V46_2",
    49: "V46_3", 50: "V46_4", 51: "V46_5", 52: "V46_6", 53: "V46_7",
    54: "V46_8", 55: "V47", 56: "V48", 57: "V49", 58: "V50", 59: "V51",
    60: "V52", 61: "V53", 62: "V53_1", 63: "V53_2", 64: "V53_3",
    65: "V53_4", 66: "V53_5", 67: "V53_6", 68: "V53_7", 69: "V53_8",
    70: "V53_9", 71: "V54", 72: "V55", 73: "V56", 74: "V57", 75: "V58",
    76: "V59", 77: "V60", 78: "V61", 79: "V62", 80: "V63", 81: "V64",
    82: "V65", 83: "V66", 84: "V66_1", 85: "V66_2", 86: "V66_3",
    87: "V66_4", 88: "V66_5", 89: "V66_6", 90: "V66_7", 91: "V66_8",
    92: "V66_9", 93: "V67", 94: "V68", 95: "V69", 96: "V70", 97: "V71",
    98: "V72", 99: "V73", 100: "V74", 101: "V75", 102: "V76", 103: "V77",
    104: "V78", 105: "V79", 106: "V80", 107: "V81", 108: "V82",
    109: "V83", 110: "V84", 111: "V85", 112: "V86", 113: "V87",
    114: "V88", 115: "V89", 116: "V90", 117: "V91", 118: "V92",
    119: "V93", 120: "V94", 121: "V95", 122: "V96", 123: "V97",
    124: "V98", 125: "V99", 126: "V100", 127: "V101", 128: "V102",
    129: "V103", 130: "V104", 131: "V105", 132: "V106", 133: "V107",
    134: "V108", 135: "V109", 136: "V110", 137: "V111", 138: "V112",
    139: "V113", 140: "V114", 141: "V114_1", 142: "V114_2", 143: "V114_3",
    144: "V114_4", 145: "V114_5", 146: "V114_6", 147: "V115", 148: "V116",
    149: "V117", 150: "V118", 151: "V119", 152: "V120", 153: "V121",
    154: "V122", 155: "V123", 156: "V124", 157: "V125", 158: "V126",
    159: "V127", 160: "V128", 161: "V129", 162: "V130", 163: "V131",
    164: "V132", 165: "V133", 166: "V134",
}

# Mapa inverso
VAR_A_POSICION = {v: k for k, v in POSICION_A_VAR.items()}

# Importar el mapa de V → ruta JSON
from scripts.generar_catalogos import VAR_MAP


def _fila_a_dict(fila: pd.Series, columnas: List[str]) -> dict:
    """Convierte una fila del DataFrame en el dict plano {VXX: valor}."""
    resultado = {}
    for col, val in zip(columnas, fila):
        if pd.isna(val) or str(val).strip() == "":
            continue
        resultado[col] = str(val).strip()
    return resultado


def _flat_a_nested(flat: dict) -> dict:
    """
    Convierte {VXX: valor} en el dict anidado que espera el validador.
    Ej: {'V7': '1985-03-15', 'V17': 'C50X'} →
        {'paciente': {'fecha_nacimiento': '1985-03-15'}, 'diagnostico': {...}}
    """
    nested: dict = {}
    for var, valor in flat.items():
        path = VAR_MAP.get(var)
        if not path:
            continue
        parts = path.split(".")
        obj = nested
        for i, p in enumerate(parts[:-1]):
            obj = obj.setdefault(p, {})
        obj[parts[-1]] = valor
    return nested


def _leer_excel(contenido: bytes) -> Tuple[pd.DataFrame, List[str]]:
    """Lee el Excel y devuelve (df, nombres_columnas_VXX)."""
    df_raw = pd.read_excel(io.BytesIO(contenido), header=None, dtype=str)

    # Detectar si la primera fila es cabecera (contiene texto no numérico)
    primera = df_raw.iloc[0].astype(str)
    es_cabecera = primera.str.contains(r'[a-zA-Z]', regex=True).any()
    if es_cabecera:
        df_raw = df_raw.iloc[1:].reset_index(drop=True)

    # Asignar nombres de columna según posición
    n_cols = len(df_raw.columns)
    col_names = [POSICION_A_VAR.get(i + 1, f"COL{i+1}") for i in range(n_cols)]
    df_raw.columns = col_names
    return df_raw, col_names


def _leer_txt(contenido: bytes) -> Tuple[pd.DataFrame, List[str]]:
    """Lee el TXT delimitado por pipes."""
    texto = contenido.decode("latin-1", errors="replace")
    lineas = [l.strip() for l in texto.splitlines() if l.strip()]
    filas = [l.split("|") for l in lineas]
    max_cols = max(len(f) for f in filas)
    col_names = [POSICION_A_VAR.get(i + 1, f"COL{i+1}") for i in range(max_cols)]
    df = pd.DataFrame(filas, columns=col_names)
    return df, col_names


def _generar_plantilla() -> bytes:
    """Genera el Excel plantilla vacío con las 166 columnas VXX como cabecera."""
    cols = [POSICION_A_VAR[i] for i in sorted(POSICION_A_VAR.keys())]
    df = pd.DataFrame(columns=cols)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Malla_CAC")
        ws = w.sheets["Malla_CAC"]
        # Ancho mínimo de columna
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 12
    buf.seek(0)
    return buf.read()


def validar_malla(
    contenido: bytes,
    nombre_archivo: str,
) -> Tuple[bytes, dict]:
    """
    Valida una malla CAC (Excel o TXT).

    Devuelve:
      - bytes del Excel de reporte de errores
      - dict resumen {total_filas, filas_con_error, filas_limpias, errores_por_variable}
    """
    from app.validators import ejecutar_validaciones
    from app.schemas.cac import CACReport
    from pydantic import ValidationError

    # Leer archivo
    ext = Path(nombre_archivo).suffix.lower()
    if ext in (".xlsx", ".xls"):
        df, columnas = _leer_excel(contenido)
    elif ext in (".txt", ".csv"):
        df, columnas = _leer_txt(contenido)
    else:
        raise ValueError(f"Formato no soportado: {ext}. Use .xlsx, .txt o .csv")

    total_filas = len(df)
    errores_lista = []
    filas_con_error = set()
    errores_por_var: dict[str, int] = {}

    for idx, fila in df.iterrows():
        num_fila = idx + 2  # +2 porque Excel empieza en 1 y tiene cabecera
        flat = _fila_a_dict(fila, columnas)
        nested = _flat_a_nested(flat)

        # Añadir cabecera mínima para el schema
        nested.setdefault("cabecera", {})
        nested["cabecera"].setdefault("id_reporte", f"FILA_{num_fila}")

        try:
            reporte = CACReport(**nested)
            resultado = ejecutar_validaciones(reporte)
            for err in resultado.errores:
                errores_lista.append({
                    "Fila": num_fila,
                    "Variable": err.variable_res or "",
                    "Campo": err.campo or "",
                    "Valor ingresado": flat.get(err.variable_res.replace("V", "V") if err.variable_res else "", ""),
                    "Código regla": err.id_regla or "",
                    "Nivel": err.nivel or "ERROR",
                    "Mensaje de error": err.mensaje or "",
                })
                filas_con_error.add(num_fila)
                var_key = err.variable_res or "GENERAL"
                errores_por_var[var_key] = errores_por_var.get(var_key, 0) + 1

        except ValidationError as ve:
            for e in ve.errors():
                campo_path = ".".join(str(x) for x in e["loc"])
                errores_lista.append({
                    "Fila": num_fila,
                    "Variable": "",
                    "Campo": campo_path,
                    "Valor ingresado": "",
                    "Código regla": "SCHEMA",
                    "Nivel": "ERROR",
                    "Mensaje de error": e["msg"],
                })
                filas_con_error.add(num_fila)

        except Exception as ex:
            errores_lista.append({
                "Fila": num_fila,
                "Variable": "",
                "Campo": "",
                "Valor ingresado": "",
                "Código regla": "PARSE",
                "Nivel": "ERROR",
                "Mensaje de error": str(ex),
            })
            filas_con_error.add(num_fila)

    # Generar Excel de errores
    df_errores = pd.DataFrame(errores_lista) if errores_lista else pd.DataFrame(
        columns=["Fila", "Variable", "Campo", "Valor ingresado", "Código regla", "Nivel", "Mensaje de error"]
    )

    # Hoja de resumen
    top_errores = sorted(errores_por_var.items(), key=lambda x: -x[1])[:20]
    df_resumen = pd.DataFrame([
        {"Métrica": "Total filas", "Valor": total_filas},
        {"Métrica": "Filas con error", "Valor": len(filas_con_error)},
        {"Métrica": "Filas limpias", "Valor": total_filas - len(filas_con_error)},
        {"Métrica": "Total errores encontrados", "Valor": len(errores_lista)},
        {"Métrica": "% filas limpias", "Valor": f"{((total_filas - len(filas_con_error)) / max(total_filas, 1) * 100):.1f}%"},
    ])
    df_top = pd.DataFrame(top_errores, columns=["Variable", "Num errores"])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df_errores.to_excel(w, index=False, sheet_name="Errores")
        df_resumen.to_excel(w, index=False, sheet_name="Resumen")
        df_top.to_excel(w, index=False, sheet_name="Top errores por variable")

        # Formato básico
        from openpyxl.styles import PatternFill, Font
        ws = w.sheets["Errores"]
        red_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
        for row in ws.iter_rows(min_row=2):
            nivel = row[5].value if len(row) > 5 else ""
            if nivel == "ERROR":
                for cell in row:
                    cell.fill = red_fill
        for col in ws.columns:
            max_w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 60)

    buf.seek(0)
    resumen = {
        "total_filas": total_filas,
        "filas_con_error": len(filas_con_error),
        "filas_limpias": total_filas - len(filas_con_error),
        "total_errores": len(errores_lista),
        "porcentaje_limpio": round((total_filas - len(filas_con_error)) / max(total_filas, 1) * 100, 1),
        "top_variables_con_error": dict(top_errores[:10]),
    }
    return buf.read(), resumen
